# Import necessary libraries
from PyQt5.QtGui import QIntValidator  # Importing integer validator from PyQt5
from PyQt5.QtWidgets import (QApplication, QPushButton, QGridLayout, QLabel, QLineEdit,  # Importing necessary widgets
                             QHBoxLayout, QVBoxLayout, QTableWidget, QTableWidgetItem, QWidget)
import sys  # Importing sys module for system-related functionality
import openpyxl  # Importing openpyxl for working with Excel files


# Define a class named Main, which inherits from QWidget
class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()  # Initialize the parent class
        # Set window geometry, size, and title
        self.setGeometry(500, 500, 500, 300)
        self.setWindowTitle("Data Collector")
        # Set tooltip for the window
        self.setToolTip("Data")

        # Create QLabel for input instruction
        self.input_lbl = QLabel(self)
        self.input_lbl.setText("Give us data...")

        # Create QLineEdit for input
        self.input_box = QLineEdit(self)

        # Create QPushButton for saving input data
        save_btn = QPushButton()
        save_btn.setText("Save")
        save_btn.clicked.connect(self.on_click)

        # Create QLineEdit for row deletion
        self.del_input_box = QLineEdit(self)
        int_validator = QIntValidator()  # Validator for integer input
        self.del_input_box.setValidator(int_validator)

        # Create QPushButton for deleting row
        del_btn = QPushButton()
        del_btn.setText("Delete Row")
        del_btn.clicked.connect(self.on_click_del)

        # Create grid layout for arranging widgets
        grid = QGridLayout()

        # Create vertical layout for table widget
        layout = QVBoxLayout()
        Vlayout = QVBoxLayout()
        self.setLayout(grid)

        # Create QTableWidget for displaying data
        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)
        Vlayout.addWidget(self.input_lbl)
        Vlayout.addWidget(self.input_box)
        Vlayout.addWidget(save_btn)
        Vlayout.addWidget(self.del_input_box)
        Vlayout.addWidget(del_btn)

        # Add layouts to the grid
        grid.addLayout(layout, 0, 0)
        grid.addLayout(Vlayout, 0, 1)

        # Load data into the table
        self.load_data()

    # Method to handle row deletion
    def on_click_del(self):
        self.delete_row(int(self.del_input_box.text()))  # Call delete_row method
        self.load_data()  # Reload data into the table

    # Static method for deleting a row from Excel
    @staticmethod
    def delete_row(num):
        wb = openpyxl.load_workbook("Data.xlsx")  # Load Excel workbook
        sheet = wb.active  # Get active sheet

        sheet.delete_rows(int(num) + 1)  # Delete row
        wb.save("Data.xlsx")  # Save changes to Excel file

    # Method to handle saving input data
    def on_click(self):
        self.add_row()  # Call add_row method
        self.input_box.clear()  # Clear input box
        self.load_data()  # Reload data into the table

    # Method to add a row of data
    def add_row(self):
        wb = openpyxl.load_workbook("Data.xlsx")  # Load Excel workbook
        sheet = wb.active  # Get active sheet

        max_rows = sheet.max_row  # Get maximum number of rows

        found = False  # Flag for indicating if input data is found

        # Loop through each row until the next row is empty
        for row_num in range(1, max_rows + 1):
            cell_value = sheet.cell(row=row_num, column=1).value
            if cell_value == self.input_box.text():  # Check if input data is found
                found = True
                # If found, increment the value in the next column
                current_count = sheet.cell(row=row_num, column=2).value or 0
                sheet.cell(row=row_num, column=2).value = current_count + 1
                break

        if not found:
            # If the search string wasn't found, add it to the next empty row with count 1
            sheet.cell(row=max_rows + 1, column=1).value = self.input_box.text()
            sheet.cell(row=max_rows + 1, column=2).value = 1

        # Save the changes to the Excel file
        wb.save("Data.xlsx")

    # Method to load data into the table
    def load_data(self):
        path = "Data.xlsx"  # Path to Excel file
        workbook = openpyxl.load_workbook(path)  # Load workbook
        sheet = workbook.worksheets[0]  # Get first worksheet

        # Set size of the table
        self.table_widget.setRowCount(sheet.max_row - 1)
        self.table_widget.setColumnCount(sheet.max_column)

        list_values = list(sheet.values)  # Get values from the sheet

        # Set headers for the table
        self.table_widget.setHorizontalHeaderLabels(list_values[0])

        row_index = 0
        # Populate the table with data
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.table_widget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1


# Main block
if __name__ == "__main__":
    app = QApplication(sys.argv)  # Create application instance
    window = Main()  # Create instance of Main class
    window.show()  # Show the window
    app.exec_()  # Execute the application